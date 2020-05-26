# !/Library/Frameworks/Python.framework/Versions/3.7/bin/python3
# -*- coding:utf-8 -*-
# @Author : Jiazhixiang
# @Time : 2019/12/11
import openpyxl
import datetime

# wb = openpyxl.Workbook()  # 利用openpyxl.Workbook()函数创建新的workbook（工作薄）对象，就是创建新的空的Excel文件。
# ws = wb.active
# ws.title = 'new title'
# ws['A1'] = 42
# ws.append([1, 2, 3])
# ws['A2'] = datetime.datetime.now()

# row = ['美国队长', '钢铁侠', '蜘蛛侠']
# ws.append(row)

# rows = [['美国队长', '钢铁侠', '蜘蛛侠'], ['是', '漫威', '宇宙', '经典', '人物']]
# for i in rows:
#     ws.append(i)
# wb.save("sample.xlsx")

"""
wb = openpyxl.Workbook()  # 利用openpyxl.Workbook()函数创建新的workbook（工作薄）对象，就是创建新的空的Excel文件。
ws = wb.active
ws.title = 'new title'
for i in range(1, 12):
    ws['A6'] = 'pitch(Hz)' + str(i)
    ws['A7'] = 'pitch(Hz)' + str(i)
    ws['A8'] = 'pitch(Hz)' + str(i)
    ws['A9'] = 'pitch(Hz)' + str(i)
    ws['A10'] = 'pitch(Hz)' + str(i)
    ws['A11'] = 'pitch(Hz)' + str(i)
    ws['A12'] = 'pitch(Hz)' + str(i)

ws.append([1, 2, 3])
ws['A2'] = datetime.datetime.now()
"""


import openpyxl

# 引用openpyxl
wb = openpyxl.Workbook()
# 利用openpyxl.Workbook()函数创建新的workbook（工作薄）对象，就是创建新的空的Excel文件。
sheet = wb.active
# wb.active就是获取这个工作薄的活动表，通常就是第一个工作簿
sheet.title = '我是一个表格'
# 可以用.title给工作表重命名。现在第一个工作表的名称就会由原来默认的“sheet1”改为"kaikeba"。
sheet['A1'] = 'hello'
# 向单个单元格写入数据
score1 = ['math', 95]
# sheet.append(score1)
# 写入整行的数据，变量类型是一个列表
wb.save('score.xlsx')
# 保存修改的Excel
wb.close()
# 关闭Excel
